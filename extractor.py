import os
import sys
import io
import requests as req
from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin
import tempfile

def get_html(url):
    print(f"[*] Fetching dynamic data from: {url} (Waiting for JS to load...)")
    
    # Method 1: Try Playwright (for JS-heavy/dynamic sites)
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                viewport={"width": 1280, "height": 800},
                ignore_https_errors=True,
                java_script_enabled=True,
            )
            page = context.new_page()
            # Try networkidle first, fallback to domcontentloaded
            try:
                page.goto(url, wait_until="networkidle", timeout=30000)
            except:
                page.goto(url, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(3000)
            html = page.content()
            browser.close()
            print("[+] Playwright fetch successful.")
            return html
    except Exception as e:
        print(f"[!] Playwright failed ({e}), trying simple HTTP fetch...")

    # Method 2: Fallback to simple requests (for static sites)
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
        }
        response = req.get(url, headers=headers, timeout=15, verify=False)
        response.raise_for_status()
        print("[+] Simple HTTP fetch successful.")
        return response.text
    except Exception as e:
        print(f"[-] Both methods failed: {e}")
        return None


def download_image_to_temp(img_url):
    """Download an image and save to temp file. Returns path or None."""
    try:
        headers = {'User-Agent': 'Mozilla/5.0 Chrome/120.0.0.0'}
        r = req.get(img_url, headers=headers, timeout=8, stream=True)
        if r.status_code == 200 and 'image' in r.headers.get('Content-Type', ''):
            suffix = '.jpg'
            ct = r.headers.get('Content-Type', '')
            if 'png' in ct: suffix = '.png'
            elif 'gif' in ct: suffix = '.gif'
            elif 'webp' in ct: suffix = '.webp'
            elif 'svg' in ct: return None  # SVG not supported by openpyxl
            
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
            for chunk in r.iter_content(1024):
                tmp.write(chunk)
            tmp.close()
            return tmp.name
    except:
        pass
    return None

def extract_data(html, base_url):
    soup = BeautifulSoup(html, 'html.parser')
    extracted_data = []
    print("[*] Parsing HTML elements...")

    # 1. Menus (Links inside <nav> or header)
    nav_tags = soup.find_all(['nav', 'header'])
    for nav in nav_tags:
        links = nav.find_all('a')
        for a in links:
            text = a.get_text(strip=True)
            href = a.get('href', '')
            if text:
                extracted_data.append({
                    "Element Type": "Menu Link",
                    "Content": text,
                    "Extra Details": urljoin(base_url, href),
                    "_img_url": None
                })

    # 2. Images & Logos
    for img in soup.find_all('img'):
        src = img.get('src', '') or img.get('data-src', '') or img.get('data-lazy', '')
        if not src:
            continue
        alt = img.get('alt', '')
        classes = " ".join(img.get('class', [])).lower()
        img_id = img.get('id', '').lower()
        full_src = urljoin(base_url, src)

        if 'logo' in src.lower() or 'logo' in alt.lower() or 'logo' in classes or 'logo' in img_id:
            el_type = "Logo"
        elif any(kw in classes for kw in ['slide', 'banner', 'hero', 'carousel']):
            el_type = "Slider / Banner Image"
        else:
            el_type = "Image"

        extracted_data.append({
            "Element Type": el_type,
            "Content": full_src,
            "Extra Details": f"Alt: {alt}",
            "_img_url": full_src
        })

    # 3. Headings (h1 - h6)
    for level in range(1, 7):
        for h in soup.find_all(f'h{level}'):
            text = h.get_text(strip=True)
            if text:
                extracted_data.append({
                    "Element Type": f"Heading (H{level})",
                    "Content": text,
                    "Extra Details": "",
                    "_img_url": None
                })

    # 4. Paragraphs
    for p in soup.find_all('p'):
        text = p.get_text(strip=True)
        if text:
            extracted_data.append({
                "Element Type": "Paragraph Text",
                "Content": text,
                "Extra Details": "",
                "_img_url": None
            })

    # 5. General Links (not in nav)
    for a in soup.find_all('a'):
        text = a.get_text(strip=True)
        href = a.get('href', '')
        if text and not a.find_parent(['nav', 'header']):
            extracted_data.append({
                "Element Type": "General Link",
                "Content": text,
                "Extra Details": urljoin(base_url, href),
                "_img_url": None
            })

    # Remove duplicates
    seen = set()
    unique_data = []
    for item in extracted_data:
        tup = (item["Element Type"], item["Content"], item["Extra Details"])
        if tup not in seen:
            seen.add(tup)
            unique_data.append(item)

    return unique_data

def save_to_excel(data, output_path):
    if not data:
        print("[-] No data extracted.")
        return

    print(f"[*] Exporting {len(data)} items to Excel (with images)...")

    # Write base data (without _img_url column)
    rows_for_df = [{"Element Type": d["Element Type"], "Content": d["Content"], "Extra Details": d["Extra Details"]} for d in data]
    df = pd.DataFrame(rows_for_df)
    df["Image Thumbnail"] = ""  # placeholder column D
    df.to_excel(output_path, index=False, engine='openpyxl')

    # Now embed images using openpyxl
    wb = load_workbook(output_path)
    ws = wb.active

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 22

    img_count = 0
    temp_files = []

    for row_idx, item in enumerate(data, start=2):  # row 1 is header
        img_url = item.get("_img_url")
        if img_url:
            tmp_path = download_image_to_temp(img_url)
            if tmp_path:
                try:
                    xl_img = XLImage(tmp_path)
                    # Scale image to thumbnail size (max 120x80)
                    max_w, max_h = 120, 80
                    if xl_img.width > max_w:
                        scale = max_w / xl_img.width
                        xl_img.width = max_w
                        xl_img.height = int(xl_img.height * scale)
                    if xl_img.height > max_h:
                        scale = max_h / xl_img.height
                        xl_img.height = max_h
                        xl_img.width = int(xl_img.width * scale)

                    cell_ref = f"D{row_idx}"
                    ws.row_dimensions[row_idx].height = 65
                    ws[cell_ref] = ""
                    ws.add_image(xl_img, cell_ref)
                    img_count += 1
                    temp_files.append(tmp_path)
                except Exception as e:
                    print(f"  [!] Could not embed image at row {row_idx}: {e}")

    wb.save(output_path)

    # Cleanup temp files
    for f in temp_files:
        try:
            os.unlink(f)
        except:
            pass

    print(f"[+] Done! {img_count} images embedded. Saved to: {os.path.abspath(output_path)}")

def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    print("="*50)
    print("   [+] Web to Excel Data Extractor")
    print("="*50)

    url = input("Enter the Website URL (e.g. https://example.com): ").strip()
    if not url:
        print("[-] Invalid URL.")
        return

    html = get_html(url)
    if html:
        data = extract_data(html, url)
        output_file = "website_data.xlsx"
        save_to_excel(data, output_file)

if __name__ == "__main__":
    main()
